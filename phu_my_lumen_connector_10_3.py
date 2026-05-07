"""
PHÚ MỸ LUMEN CONNECTOR - v4.0
Cài đặt: pip install streamlit pandas openpyxl python-dateutil
Chạy   : streamlit run phu_my_lumen_connector_v4.py

Thay đổi v4.0 (so với v3.8):
  FIX 1: Bỏ bảng xem trước kết quả (nặng, không cần thiết) → xuất file theo dõi để xem bằng Excel
  FIX 2: Sửa dedup bug trong NHIEU_MA section: key cũ (MA, dd) → key mới (MA, dd, ten)
          → Morphin 40.16 + 40.43 giờ hiển thị đúng cả 2 dạng TEN trong dropdown
  FIX 3: Insulin analog 40.805 / 40.805.1 / 40.30.805.2 hiển thị đủ tùy theo TEN file thầu
  FIX 4: Magnesi aspartat: _norm_tight chuẩn hóa dấu cách quanh + nhất quán → khớp đúng TEN
          → TEN_HOAT_CHAT_XK lấy từ file Tân dược, không tự sửa từ file thầu
  FIX 5: Bỏ bảng mã đường dùng chuẩn (không cần thiết, làm nặng UI)
  FIX 6: Tối ưu hiệu năng: cache @st.cache_data cho build_thuoc_lookup & build_dd_lookup
          → không re-build mỗi lần bấm nút
  FIX 7: Lưu trữ nhà thầu qua st.session_state + download/upload CSV alias để persist
          khi deploy trên Streamlit Cloud (không dùng file local nữa)
"""

import csv
import io
import os
import re
import tempfile
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import streamlit as st

# ================================================================
# MODULE ALIAS (chỉ dùng session_state, không ghi file local)
# ================================================================
ALIAS_COLS    = ['TEN_THAU', 'TEN_TANDUC', 'GHI_CHU']
ALIAS_DD_COLS = ['TEN_HOAT_CHAT', 'DD_GOC', 'DD_CHUAN', 'GHI_CHU']

def _empty_alias():    return pd.DataFrame(columns=ALIAS_COLS)
def _empty_alias_dd(): return pd.DataFrame(columns=ALIAS_DD_COLS)

def build_alias_lookup(df_alias: pd.DataFrame) -> dict:
    lk = {}
    for _, r in df_alias.iterrows():
        k = _norm(r.get('TEN_THAU', ''))
        v = r.get('TEN_TANDUC', '').strip()
        if k and v:
            lk[k] = v
    return lk

def build_alias_dd_lookup(df_dd: pd.DataFrame) -> dict:
    lk = {}
    for _, r in df_dd.iterrows():
        hc   = _norm(r.get('TEN_HOAT_CHAT', ''))
        dd_g = _norm(r.get('DD_GOC', ''))
        dd_c = r.get('DD_CHUAN', '').strip()
        if hc and dd_g and dd_c:
            lk[(hc, dd_g)] = dd_c
    return lk

# ================================================================
# MODULE TIỀN XỬ LÝ
# ================================================================
def sc(val) -> str:
    if pd.isna(val): return ''
    return str(val).replace('\n', ' ').replace('\r', '').strip()

def sanitize_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [sc(c) for c in df.columns]
    for col in df.columns:
        df[col] = df[col].apply(sc)
    return df

def find_header(df_raw, keywords=('STT', 'Tên hoạt chất', 'Số lượng', 'Nhà thầu')) -> int:
    for i, row in df_raw.iterrows():
        text = ' '.join(str(v) for v in row.values if str(v) != 'nan')
        if sum(1 for kw in keywords if kw in text) >= 2:
            return i
    return 0

def read_excel(file_bytes: bytes, keywords=None) -> pd.DataFrame:
    kw = keywords or ('STT', 'Tên hoạt chất', 'Số lượng')
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp.write(file_bytes); path = tmp.name
    df_raw = pd.read_excel(path, header=None, dtype=str)
    hrow   = find_header(df_raw, kw)
    df     = pd.read_excel(path, header=hrow, dtype=str)
    os.unlink(path)
    return sanitize_df(df).dropna(how='all').reset_index(drop=True)

def read_xls_compat(file_bytes: bytes) -> pd.DataFrame:
    buf = io.BytesIO(file_bytes)
    for engine in ('openpyxl', None):
        buf.seek(0)
        try:
            kw = {'engine': engine} if engine else {}
            return sanitize_df(pd.read_excel(buf, header=0, dtype=str, **kw))
        except Exception: pass
    raise RuntimeError("Không đọc được file. Thử: pip install xlrd==1.2.0")

# ================================================================
# MODULE REGEX - TÁCH SĐK
# ================================================================
RE_NEW    = re.compile(r'\b(\d{12})\b')
RE_OLD_VN = re.compile(r'\b(VN\d*-\d+-\d+)\b',    re.IGNORECASE)
RE_OLD_VD = re.compile(r'\b(VD-\d+-\d+)\b',        re.IGNORECASE)
RE_OLD_QL = re.compile(r'\b((?:QLSP|QLDB|QLĐB|GPNK|GC|DP|PB|TN|SP\d+)-[\d\w/-]+)\b', re.IGNORECASE)
RE_QLD    = re.compile(r'(\d{3,5}/QLD-\w+)', re.IGNORECASE)

def _csdk(s): return s.strip().rstrip('.,;) ')

def parse_sdk(sdk_raw: str) -> list:
    if not sdk_raw: return [('', True, False)]
    sdk_raw = str(sdk_raw).strip().strip('"')
    new_list = RE_NEW.findall(sdk_raw)
    old_list, seen = [], set()
    for pat in (RE_OLD_VN, RE_OLD_VD, RE_OLD_QL):
        for m in pat.findall(sdk_raw):
            v = _csdk(m)
            if v and v not in seen: seen.add(v); old_list.append(v)
    qld_list = [x.strip() for x in RE_QLD.findall(sdk_raw)]
    if new_list and old_list:
        return [(_csdk(new_list[0]), True, False)] + [(o, False, False) for o in old_list]
    if new_list: return [(_csdk(new_list[0]), True, False)]
    if old_list: return [(old_list[0], True, False)]
    if qld_list:
        return [(qld_list[0], True, True)] + [(q, False, True) for q in qld_list[1:]]
    cleaned = re.split(r'[\.(]', sdk_raw)[0].strip()
    return [(cleaned or sdk_raw, True, False)]

def expand_sdk_rows(df, col_sdk, col_sl):
    rows = []
    for _, row in df.iterrows():
        for sdk_val, is_primary, is_qld in parse_sdk(str(row.get(col_sdk, ''))):
            r = row.copy()
            r[col_sdk] = sdk_val
            r['IS_QLD_KD'] = 'CẦN RÀ SOÁT' if is_qld else ''
            if not is_primary: r[col_sl] = '0'
            rows.append(r)
    return pd.DataFrame(rows).reset_index(drop=True)

# ================================================================
# MODULE CHUẨN HÓA & TRA CỨU
# ================================================================
def _norm(s: str) -> str:
    s = str(s).lower().strip()
    s = re.sub(r'\s*\+\s*', ' + ', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()

def _norm_tight(s: str) -> str:
    s = str(s).lower().strip()
    s = re.sub(r'\s*\+\s*', '+', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()

def _count_components(s: str) -> int:
    return len(re.split(r'\s*\+\s*', s.strip()))

_DD_ALIAS_BUILTIN: list = [
    (['tiêm bắp (im)', 'tiêm bắp im', 'tiêm bắp', 'im', 'intramuscular'], 'tiêm bắp'),
    (['tiêm tĩnh mạch (iv)', 'tiêm tĩnh mạch iv', 'tiêm tĩnh mạch', 'iv', 'intravenous', 'tiêm tm'], 'tiêm tĩnh mạch'),
    (['tiêm dưới da (sc)', 'tiêm dưới da sc', 'tiêm dưới da', 'sc', 'subcutaneous'], 'tiêm dưới da'),
    (['uống', 'viên nang', 'viên nén', 'oral', 'per os'], 'uống'),
    (['nhỏ mắt', 'tra mắt', 'ophthalmic'], 'nhỏ mắt'),
    (['dùng ngoài da', 'bôi ngoài da', 'bôi da', 'topical'], 'dùng ngoài da'),
    (['đặt âm đạo', 'đặt âm đạo (vaginal)', 'vaginal'], 'đặt âm đạo'),
    (['đặt hậu môn', 'rectal', 'đặt trực tràng'], 'đặt hậu môn'),
    (['hít', 'xịt', 'inhalation', 'inhaled'], 'hít'),
    (['truyền tĩnh mạch', 'truyền tm', 'iv infusion', 'infusion'], 'truyền tĩnh mạch'),
    (['màng bụng', 'dùng theo đường màng bụng', 'lọc màng bụng', 'thẩm phân', 'peritoneal', 'dialysis'], 'dùng theo đường màng bụng'),
]

def _dd_canonical(dd: str) -> str:
    n = _norm(dd)
    for variants, canonical in _DD_ALIAS_BUILTIN:
        for v in variants:
            if v in n or n in v:
                return canonical
    return n

def _dd_similarity(dd_thau: str, dd_td: str) -> float:
    import difflib as _dl
    n1 = _norm_tight(dd_thau)
    n2 = _norm_tight(dd_td)
    if n1 == n2: return 1.0
    if n1 == 'tiêm' and n2.startswith('tiêm'): return 0.9
    if n1 == 'tiem' and n2.startswith('tiem'): return 0.9
    if n1 in ('tiêm truyền', 'tiem truyen') and ('truyền' in n2 or 'truyen' in n2): return 0.88
    c1 = _dd_canonical(dd_thau)
    c2 = _dd_canonical(dd_td)
    if c1 and c2 and c1 == c2: return 0.85
    n1f = _norm(dd_thau); n2f = _norm(dd_td)
    ratio = _dl.SequenceMatcher(None, n1f, n2f).ratio()
    words1 = set(n1f.split()); words2 = set(n2f.split())
    common = words1 & words2
    if common:
        ratio = min(0.8, ratio + 0.1 * len(common))
    return ratio

@st.cache_data(show_spinner=False)
def build_thuoc_lookup(thuoc_bytes: bytes):
    """
    Cache by file bytes hash. Xây dựng từ điển tra cứu MA_THUOC.

    FIX v4.0: goi_y dedup dùng (MA, dd_tight, ten_tight) thay vì (MA, dd_tight)
    → Morphin 40.16 với TEN="Morphin" và TEN="Morphin (hydroclorid, sulfat)"
      cùng DD="Tiêm" đều được giữ riêng biệt.
    """
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp.write(thuoc_bytes); path = tmp.name
    df_thuoc = sanitize_df(pd.read_excel(path, header=0, dtype=str))
    os.unlink(path)

    lk           = {}
    name_map     = {}
    goi_y        = {}
    conflict_map = {}

    for _, r in df_thuoc.iterrows():
        ten_raw = sc(r.get('TEN', ''))
        dd_raw  = sc(r.get('DUONGDUNG', ''))
        ma      = sc(r.get('MA', ''))
        ten_t   = _norm_tight(ten_raw)
        dd_t    = _norm_tight(dd_raw)
        if not ten_t or not ma:
            continue

        key = (ten_t, dd_t)
        conflict_map.setdefault(key, set())
        conflict_map[key].add(ma)

        if key not in lk:
            lk[key]       = ma
            name_map[key] = ten_raw

        # FIX: dedup dùng (MA, dd_tight, ten_tight) → giữ cả Morphin và Morphin (hydroclorid, sulfat)
        goi_y.setdefault(ten_t, {})
        pair_k = f"{ma}||{dd_t}||{ten_t}"   # KEY MỚI: thêm ten_tight
        if pair_k not in goi_y[ten_t]:
            goi_y[ten_t][pair_k] = (dd_raw, ma, ten_raw)

    goi_y_list = {
        hc: [(dd, ma, ten) for dd, ma, ten in ma_dd.values()]
        for hc, ma_dd in goi_y.items()
    }
    conflict_map = {k: sorted(v) for k, v in conflict_map.items() if len(v) > 1}

    # Trả về df_thuoc để dùng ở các chỗ khác
    return lk, name_map, goi_y_list, conflict_map, df_thuoc

@st.cache_data(show_spinner=False)
def build_dd_lookup_cached(cach_bytes: bytes) -> dict:
    buf = io.BytesIO(cach_bytes)
    try:
        df_cach = sanitize_df(pd.read_excel(buf, header=0, dtype=str))
    except Exception:
        df_cach = read_xls_compat(cach_bytes)
    lk = {}
    cols = df_cach.columns.tolist()
    col_ma, col_ten = cols[0], cols[1]
    for _, r in df_cach.iterrows():
        ma  = sc(r.get(col_ma, ''))
        ten = _norm(sc(r.get(col_ten, '')))
        if ten and ma:
            lk[ten] = ma
    return lk

def lookup_thuoc(hoat_chat: str, duong_dung: str,
                 lk: dict, name_map: dict, goi_y: dict,
                 alias_ten_lk: dict, alias_dd_lk: dict):
    if not hoat_chat:
        return '', '', []
    n_comp_thau = _count_components(hoat_chat)
    hc_tight    = _norm_tight(hoat_chat)
    dd_tight    = _norm_tight(duong_dung) if duong_dung else ''
    dd_use      = dd_tight
    hc_candidates = [hc_tight]

    # Tầng 1: Khớp chính xác tên + đường dùng
    for k_hc in hc_candidates:
        key = (k_hc, dd_use)
        if key in lk:
            return lk[key], name_map.get(key, hoat_chat), []

    # Tầng 2: Khớp chính xác tên + canonical đường dùng
    dd_canon = _dd_canonical(dd_use)
    for k_hc in hc_candidates:
        for (lk_hc, lk_dd), ma in lk.items():
            if lk_hc == k_hc and _dd_canonical(lk_dd) == dd_canon and dd_canon:
                ten_chuan = name_map.get((lk_hc, lk_dd), hoat_chat)
                return ma, ten_chuan, []

    # Tầng 3: Hints — chỉ từ goi_y[hc_tight]
    hints_all = goi_y.get(hc_tight, [])
    hints_filtered = [e for e in hints_all if _count_components(e[2]) == n_comp_thau]
    dd_orig_n = _norm_tight(duong_dung) if duong_dung else ''
    hints_filtered.sort(key=lambda e: _dd_similarity(dd_orig_n, e[0]), reverse=True)
    return '', '', hints_filtered

def lookup_dd(duong_dung: str, lk: dict) -> str:
    if not duong_dung: return ''
    return lk.get(_norm(duong_dung), '')

# ================================================================
# MODULE NHÓM & NGÀY
# ================================================================
def convert_nhom(raw: str) -> str:
    if not raw: return ''
    m = re.search(r'[Nn]h[oóô]m\s*(\d+)', raw)
    if m: return f"N{m.group(1)}"
    m2 = re.match(r'^(N\d+)$', raw.strip(), re.IGNORECASE)
    if m2: return m2.group(1).upper()
    return raw.strip()

def to_yyyymmdd(d) -> str:
    if d is None: return ''
    if isinstance(d, (datetime, date)): return d.strftime('%Y%m%d')
    return ''

def add_months(d: date, n: int) -> date:
    return d + relativedelta(months=int(n))

# ================================================================
# MODULE XUẤT FILE MAU_03
# ================================================================
MAU03_COLS = [
    'STT','MA_THUOC','TEN_HOAT_CHAT','TEN_THUOC','DON_VI_TINH',
    'HAM_LUONG','DUONG_DUNG','MA_DUONG_DUNG','DANG_BAO_CHE',
    'SO_DANG_KY','SO_LUONG','DON_GIA','DON_GIA_BH','QUY_CACH',
    'NHA_SX','NUOC_SX','NHA_THAU','TT_THAU',
    'TU_NGAY_HD','DEN_NGAY_HD','MA_CSKCB','LOAI_THUOC',
    'LOAI_THAU','HT_THAU','MA_DVKT','TCCL','BO_PHAN_VT',
    'TEN_KHOA_HOC','NGUON_GOC','PP_CHEBIEN','MA_DL_NHAP',
    'MA_DL_CB','TLHH_CB','TLHH_BQ','MA_CSKCB_THUOC',
    'TU_NGAY','DEN_NGAY',
]

RED_FILL  = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
ORG_FILL  = PatternFill(start_color='FFE5CC', end_color='FFE5CC', fill_type='solid')
YEL_FILL  = PatternFill(start_color='FFFACC', end_color='FFFACC', fill_type='solid')
PUR_FILL  = PatternFill(start_color='E8D5FF', end_color='E8D5FF', fill_type='solid')
NORM_FONT = Font(name='Arial', size=10)

def export_mau03(df: pd.DataFrame, template_bytes: bytes, filter_nt=None) -> bytes:
    wb = load_workbook(io.BytesIO(template_bytes))
    ws = wb.active
    hmap = {}
    for c in range(1, ws.max_column + 2):
        val = ws.cell(1, c).value
        if val: hmap[str(val).strip()] = c

    df_out = df.copy()
    if filter_nt is not None:
        df_out = df_out[df_out['NHA_THAU'].isin(filter_nt)]
    df_out = df_out.reset_index(drop=True)

    for row_i, (_, row) in enumerate(df_out.iterrows(), start=2):
        no_ma    = not str(row.get('MA_THUOC', '')).strip()
        is_qld   = str(row.get('IS_QLD_KD', '')) == 'CẦN RÀ SOÁT'
        is_no_dd = not str(row.get('MA_DUONG_DUNG', '')).strip() and str(row.get('DUONG_DUNG','')).strip()
        is_sub   = str(row.get('IS_MA_CON', '')) == '1'

        for col_name, col_idx in hmap.items():
            val = row.get(col_name, '')
            val = '' if pd.isna(val) else str(val)
            if col_name == 'MA_THUOC' and val:
                cell = ws.cell(row=row_i, column=col_idx, value=val)
                cell.number_format = '@'
            else:
                cell = ws.cell(row=row_i, column=col_idx, value=val or None)
            cell.font = NORM_FONT
            if no_ma:      cell.fill = RED_FILL
            elif is_qld:   cell.fill = ORG_FILL
            elif is_no_dd: cell.fill = YEL_FILL
            elif is_sub:   cell.fill = PUR_FILL

    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()

# ================================================================
# ÁNH XẠ CỘT
# ================================================================
COL_KW = {
    'TEN_HOAT_CHAT':['hoạt chất mời thầu','hoạt chất','thành phần'],
    'TEN_THUOC':    ['tên thuốc','biệt dược'],
    'DON_VI_TINH':  ['đơn vị tính','đvt','đơn vị'],
    'HAM_LUONG':    ['nồng độ','hàm lượng'],
    'DUONG_DUNG':   ['đường dùng'],
    'DANG_BAO_CHE': ['dạng bào chế'],
    'SO_DANG_KY':   ['gđklh','gpnk','số đăng ký','định danh'],
    'SO_LUONG':     ['số lượng'],
    'DON_GIA':      ['đơn giá trúng thầu','đơn giá'],
    'QUY_CACH':     ['quy cách'],
    'NHA_SX':       ['tên cơ sở sản xuất', 'cơ sở sản xuất', 'nhà sản xuất', 'manufacturer'],
    'NUOC_SX':      ['xuất xứ', 'nước sản xuất', 'nuoc sx', 'nước sx', 'xuat xu', 'origin', 'country'],
    'NHA_THAU':     ['tên nhà thầu','nhà thầu'],
    'NHOM_THUOC':   ['nhóm thuốc','nhóm'],
    'GOI_THAU':     ['gói'],
}

_FUZZY_FRAGMENTS = {'NUOC_SX': ['xuất xứ', 'xuất', 'xứ', 'origin', 'country']}
_NUOC_SX_BLACKLIST = ['sản xuất', 'nhà sản', 'cơ sở']

def suggest_col(source_cols, target):
    kws = COL_KW.get(target, [target.lower()])
    sl  = [c.lower() for c in source_cols]
    for kw in kws:
        kw_norm = kw.lower()
        for i, sc_val in enumerate(sl):
            if kw_norm in sc_val:
                if target == 'NUOC_SX':
                    if any(bl in sc_val for bl in _NUOC_SX_BLACKLIST):
                        continue
                return source_cols[i]
    for frag in _FUZZY_FRAGMENTS.get(target, []):
        for i, sc_val in enumerate(sl):
            if frag in sc_val:
                if target == 'NUOC_SX':
                    if any(bl in sc_val for bl in _NUOC_SX_BLACKLIST):
                        continue
                return source_cols[i]
    return ''

# ================================================================
# STREAMLIT CONFIG
# ================================================================
st.set_page_config(page_title="Phú Mỹ Lumen Connector v4.0", page_icon="💊",
                   layout="wide", initial_sidebar_state="collapsed")
st.markdown("""
<style>
.step-header{background:#f0f4f8;border-left:4px solid #2e6da4;
    padding:8px 16px;border-radius:4px;margin:20px 0 8px 0;
    font-weight:600;color:#1e3a5f;}
</style>""", unsafe_allow_html=True)
st.markdown("""
<div style="background:linear-gradient(90deg,#1e3a5f,#2e6da4);
     padding:18px 24px;border-radius:10px;margin-bottom:20px;color:white">
<h2 style="margin:0">💊 PHÚ MỸ LUMEN CONNECTOR <span style="font-size:13px;opacity:.7">v4.0</span></h2>
<p style="margin:4px 0 0 0;opacity:.85;font-size:14px">
Module xử lý & ánh xạ danh mục thuốc trúng thầu → Xuất Mẫu 03 BHYT</p>
</div>""", unsafe_allow_html=True)

# Session state init
_SS_DEFAULTS = {
    'df_result': None,
    'mau03_bytes': None,
    'nha_thau_info': {},
    'df_alias': None,
    'df_alias_dd': None,
    'ma_chon_override': {},
    'ma_chon_manual': {},
}
for k, v in _SS_DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v
if st.session_state['df_alias'] is None:
    st.session_state['df_alias'] = _empty_alias()
if st.session_state['df_alias_dd'] is None:
    st.session_state['df_alias_dd'] = _empty_alias_dd()

tab_main, tab_alias = st.tabs(["🏠 Xử lý chính", "✏️ Quản lý Alias Hoạt chất"])

# ════════════════════════════════════════════════════
# TAB 1: XỬ LÝ CHÍNH
# ════════════════════════════════════════════════════
with tab_main:

    # ── BƯỚC 1: UPLOAD ──────────────────────────────
    st.markdown('<div class="step-header">📁 Bước 1 — Tải lên 4 file dữ liệu</div>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown("**📋 File Danh mục thầu**")
        file_thau  = st.file_uploader("QĐ phê duyệt (.xlsx/.xls)", type=['xlsx','xls'], key='up_thau')
    with c2:
        st.markdown("**💊 File Thuốc tân dược**")
        file_thuoc = st.file_uploader("Thuốc tân dược (.xlsx)",     type=['xlsx'],       key='up_thuoc')
    with c3:
        st.markdown("**🔬 File Cách dùng thuốc**")
        file_cach  = st.file_uploader("Cách dùng (.xls/.xlsx)",     type=['xlsx','xls'], key='up_cach')
    with c4:
        st.markdown("**📄 Template MAU_03**")
        file_mau03 = st.file_uploader("MAU_03.xlsx",                type=['xlsx'],       key='up_mau03')

    if not all([file_thau, file_thuoc, file_cach, file_mau03]):
        st.info("👆 Vui lòng tải lên đủ 4 file để bắt đầu.")
        st.stop()

    # ── BƯỚC 2: ĐỌC FILE ────────────────────────────
    st.markdown('<div class="step-header">⚙️ Bước 2 — Đọc & phân tích dữ liệu</div>', unsafe_allow_html=True)

    current_file_id = getattr(file_thau, 'file_id', None) or file_thau.name
    if st.session_state.get('_last_file_id') != current_file_id:
        st.session_state['_last_file_id'] = current_file_id
        for k in list(st.session_state.keys()):
            if k.startswith('cm_'):
                del st.session_state[k]

    with st.spinner("Đang đọc các file..."):
        thau_bytes  = file_thau.read()
        thuoc_bytes = file_thuoc.read()
        cach_bytes  = file_cach.read()
        mau03_bytes = file_mau03.read()
        st.session_state['mau03_bytes'] = mau03_bytes

        df_thau = read_excel(thau_bytes, keywords=('STT','Tên hoạt chất','Số lượng','Nhà thầu'))

        # Cache: build lookup từ bytes hash → không re-build khi bấm nút
        thuoc_lk, name_map, goi_y_dd, conflict_map, df_thuoc = build_thuoc_lookup(thuoc_bytes)
        dd_lk = build_dd_lookup_cached(cach_bytes)

    alias_ten_lk = build_alias_lookup(st.session_state['df_alias'])
    alias_dd_lk  = build_alias_dd_lookup(st.session_state['df_alias_dd'])

    r1, r2, r3, r4 = st.columns(4)
    r1.success(f"✅ File thầu: **{len(df_thau)}** dòng")
    r2.success(f"✅ Tân dược: **{len(goi_y_dd)}** hoạt chất")
    r3.success(f"✅ Đường dùng: **{len(dd_lk)}** loại")
    r4.success(f"✅ Alias tên: **{len(st.session_state['df_alias'])}** | ĐD: **{len(st.session_state['df_alias_dd'])}** cặp")

    # ── BƯỚC 3: ÁNH XẠ CỘT ─────────────────────────
    st.markdown('<div class="step-header">🗂️ Bước 3 — Ánh xạ cột file thầu → Template MAU_03</div>', unsafe_allow_html=True)
    st.caption("Hệ thống tự động gợi ý. Điều chỉnh nếu cần.")

    source_cols = df_thau.columns.tolist()
    SKIP = '-- Bỏ qua --'; opts = [SKIP] + source_cols
    MAP_TARGETS = [
        ('TEN_HOAT_CHAT','Tên hoạt chất'),('TEN_THUOC','Tên thuốc'),
        ('DON_VI_TINH','Đơn vị tính'),('HAM_LUONG','Nồng độ/Hàm lượng'),
        ('DUONG_DUNG','Đường dùng'),('DANG_BAO_CHE','Dạng bào chế'),
        ('SO_DANG_KY','Số đăng ký (GĐKLH/GPNK)'),('SO_LUONG','Số lượng'),
        ('DON_GIA','Đơn giá trúng thầu'),('QUY_CACH','Quy cách'),
        ('NHA_SX','Tên cơ sở sản xuất'),('NUOC_SX','Xuất xứ'),
        ('NHA_THAU','Tên nhà thầu'),('NHOM_THUOC','Nhóm thuốc'),('GOI_THAU','Gói thầu'),
    ]

    _COL_PREFER = {
        'NUOC_SX': ['xuất xứ', 'nước sản xuất', 'nuoc_sx', 'origin', 'country'],
        'NHA_SX':  ['tên cơ sở sản xuất', 'cơ sở sản xuất', 'nhà sản xuất', 'manufacturer'],
    }

    def _find_col(col_list, keywords):
        for kw in keywords:
            kw_l = kw.lower().strip()
            for col in col_list:
                if kw_l in col.lower().strip():
                    return col
        return ''

    for tgt, _ in MAP_TARGETS:
        ss_key = f'cm_{tgt}'
        if ss_key not in st.session_state:
            prefer_kws = _COL_PREFER.get(tgt)
            sug = _find_col(source_cols, prefer_kws) if prefer_kws else suggest_col(source_cols, tgt)
            st.session_state[ss_key] = sug if sug else SKIP

    col_map = {}
    grid = st.columns(4)
    for i, (tgt, label) in enumerate(MAP_TARGETS):
        with grid[i % 4]:
            chosen = st.selectbox(f"`{tgt}` ← {label}", opts, key=f'cm_{tgt}')
            if chosen != SKIP:
                col_map[tgt] = chosen

    # ── BƯỚC 4: CẤU HÌNH THẦU ───────────────────────
    st.markdown('<div class="step-header">📝 Bước 4 — Thông tin thầu chung</div>', unsafe_allow_html=True)
    cc1, cc2, cc3, cc4 = st.columns(4)
    with cc1:
        so_qd    = st.text_input("Số quyết định", value="449/QĐ-BVĐN")
        ma_cskcb = st.text_input("Mã CSKCB", value="48001")
    with cc2:
        nam_thau     = st.text_input("Năm thầu", value="2026")
        goi_mac_dinh = st.text_input("Gói thầu mặc định", value="G1")
    with cc3:
        loai_thuoc = st.text_input("LOAI_THUOC", value="1")
        loai_thau  = st.text_input("LOAI_THAU", value="1")
    with cc4:
        ht_thau = st.text_input("HT_THAU", value="1")

    # ── BƯỚC 5: CHẠY XỬ LÝ ─────────────────────────
    st.markdown('<div class="step-header">🚀 Bước 5 — Chạy xử lý & ánh xạ dữ liệu</div>', unsafe_allow_html=True)

    if st.button("▶️  Chạy xử lý dữ liệu", type="primary", use_container_width=True):
        with st.spinner("Đang xử lý..."):
            df_work = pd.DataFrame()
            for tgt, src in col_map.items():
                if src in df_thau.columns:
                    df_work[tgt] = df_thau[src].apply(
                        lambda x: '' if str(x).strip() in ('nan','NaN','None','') else str(x))
                else:
                    df_work[tgt] = ''

            if 'TEN_HOAT_CHAT' in df_work.columns:
                df_work = df_work[df_work['TEN_HOAT_CHAT'].str.strip() != '']
            df_work = df_work.reset_index(drop=True)

            if 'SO_DANG_KY' in df_work.columns and 'SO_LUONG' in df_work.columns:
                df_work = expand_sdk_rows(df_work, 'SO_DANG_KY', 'SO_LUONG')

            df_work['STT'] = range(1, len(df_work)+1)

            result_pairs = df_work.apply(
                lambda r: lookup_thuoc(r.get('TEN_HOAT_CHAT',''), r.get('DUONG_DUNG',''),
                                       thuoc_lk, name_map, goi_y_dd,
                                       alias_ten_lk, alias_dd_lk), axis=1)
            df_work['MA_THUOC']            = result_pairs.apply(lambda x: x[0])
            df_work['TEN_HOAT_CHAT_CHUAN'] = result_pairs.apply(lambda x: x[1])
            df_work['GOI_Y_DD'] = result_pairs.apply(
                lambda x: ' | '.join([f"{d} → {m}" for d,m,_ in x[2]]) if x[2] else '')
            df_work['TEN_HOAT_CHAT_XK'] = df_work.apply(
                lambda r: r['TEN_HOAT_CHAT_CHUAN'] if r['MA_THUOC'] and r['TEN_HOAT_CHAT_CHUAN']
                          else r['TEN_HOAT_CHAT'],
                axis=1)

            dd_series = df_work.get('DUONG_DUNG', pd.Series(['']*len(df_work)))
            df_work['MA_DUONG_DUNG'] = dd_series.apply(lambda x: lookup_dd(x, dd_lk))

            df_work['IS_MA_CON'] = df_work['MA_THUOC'].apply(
                lambda m: '1' if re.search(r'\.\d+$', str(m)) and len(str(m).split('.')) >= 3 else '')

            def check_conflict(row):
                hc     = _norm_tight(row.get('TEN_HOAT_CHAT',''))
                dd     = _norm_tight(row.get('DUONG_DUNG',''))
                key    = (hc, dd)
                mas    = conflict_map.get(key, [])
                return ' | '.join(mas) if mas else ''
            df_work['NHIEU_MA'] = df_work.apply(check_conflict, axis=1)

            df_work['NHOM_MA'] = df_work.get('NHOM_THUOC', pd.Series(['']*len(df_work))).apply(convert_nhom)
            def build_tt(row):
                goi = str(row.get('GOI_THAU','')).strip() or goi_mac_dinh
                return f"{so_qd};{goi};{row.get('NHOM_MA','')};{nam_thau}"
            df_work['TT_THAU']    = df_work.apply(build_tt, axis=1)
            df_work['DON_GIA_BH'] = df_work.get('DON_GIA', pd.Series(['']*len(df_work)))
            df_work['MA_CSKCB']   = ma_cskcb
            df_work['MA_CSKCB_THUOC'] = ''
            df_work['LOAI_THUOC'] = loai_thuoc
            df_work['LOAI_THAU']  = loai_thau
            df_work['HT_THAU']    = ht_thau
            for c in ('TU_NGAY_HD','DEN_NGAY_HD','TU_NGAY','DEN_NGAY'):
                df_work[c] = ''

            st.session_state['df_result'] = df_work
            # Reset chọn mã khi chạy lại
            st.session_state['ma_chon_override'] = {}
            st.session_state['ma_chon_manual']   = {}

        n_qld  = (df_work.get('IS_QLD_KD', pd.Series()) == 'CẦN RÀ SOÁT').sum()
        n_nodd = (df_work['MA_DUONG_DUNG'] == '').sum()
        n_sub  = (df_work['IS_MA_CON'] == '1').sum()
        st.success(f"✅ Xử lý xong! Tổng **{len(df_work):,}** dòng.")
        if n_qld:  st.warning(f"🟠 **{n_qld}** dòng SĐK dạng QLD-KD — tô cam, cần rà soát thủ công.")
        if n_nodd: st.warning(f"🟡 **{n_nodd}** dòng đường dùng KHÔNG KHỚP bảng chuẩn — tô vàng.")
        if n_sub:  st.warning(f"🟣 **{n_sub}** dòng có mã dạng XX.XXX.1/.2 — tô tím, cần chọn đúng mã.")

    # ── BƯỚC 6: NHÀ THẦU & NGÀY KÝ ─────────────────
    if st.session_state.get('df_result') is not None:
        df_result = st.session_state['df_result'].copy()

        st.markdown('<div class="step-header">🏢 Bước 6 — Quản lý nhà thầu & Ngày ký phụ lục</div>', unsafe_allow_html=True)

        nha_thau_list = sorted([
            x for x in df_result['NHA_THAU'].dropna().unique()
            if str(x).strip() not in ('','nan')
        ])
        st.info(f"📋 Tìm thấy **{len(nha_thau_list)}** nhà thầu.")

        nha_thau_info: dict = st.session_state['nha_thau_info']
        for nt in nha_thau_list:
            if nt not in nha_thau_info:
                nha_thau_info[nt] = {'da_ky': False}

        # Hướng dẫn lưu trữ: download CSV
        with st.expander("💾 Lưu & Khôi phục dữ liệu nhà thầu (tránh mất khi refresh)", expanded=False):
            st.caption("Streamlit Cloud không lưu file local. Tải CSV này về máy sau mỗi phiên, upload lại khi cần.")
            # Export current state
            nt_rows = []
            for nt, info in nha_thau_info.items():
                if info.get('da_ky') and info.get('ngay_ky'):
                    nt_rows.append({
                        'NHA_THAU': nt,
                        'NGAY_KY':  info['ngay_ky'].strftime('%Y%m%d'),
                        'THOI_HAN': str(info.get('thoi_han', 12)),
                        'DEN_NGAY': info['den_ngay'].strftime('%Y%m%d') if info.get('den_ngay') else '',
                    })
            if nt_rows:
                csv_bytes = pd.DataFrame(nt_rows).to_csv(index=False).encode('utf-8')
                st.download_button("⬇️ Tải CSV nhà thầu đã ký", csv_bytes,
                                   f"nha_thau_{datetime.now().strftime('%Y%m%d')}.csv",
                                   "text/csv", use_container_width=True)
            uploaded_log = st.file_uploader("⬆️ Upload CSV nhà thầu đã lưu", type=['csv'], key='upload_log')
            if uploaded_log:
                try:
                    df_log = pd.read_csv(uploaded_log, dtype=str).fillna('')
                    n_loaded = 0
                    for _, row in df_log.iterrows():
                        nt = row.get('NHA_THAU','').strip()
                        if not nt: continue
                        try:
                            ngay_ky  = datetime.strptime(row['NGAY_KY'], '%Y%m%d').date()
                            thoi_han = int(row['THOI_HAN'])
                            den_ngay = datetime.strptime(row['DEN_NGAY'], '%Y%m%d').date() if row.get('DEN_NGAY') else add_months(ngay_ky, thoi_han)
                            nha_thau_info[nt] = {'da_ky': True, 'ngay_ky': ngay_ky, 'thoi_han': thoi_han, 'den_ngay': den_ngay}
                            n_loaded += 1
                        except Exception: pass
                    st.session_state['nha_thau_info'] = nha_thau_info
                    st.success(f"✅ Đã khôi phục {n_loaded} nhà thầu từ CSV."); st.rerun()
                except Exception as e:
                    st.error(f"❌ Lỗi: {e}")

        # Điền ngày hàng loạt
        st.subheader("⚡ Điền ngày nhanh cho nhiều nhà thầu")
        with st.expander("Mở để điền 1 ngày cho nhiều công ty", expanded=False):
            selected_bulk = st.multiselect(
                "Chọn nhà thầu cần điền ngày",
                options=nha_thau_list, default=[], key='bulk_select')
            b1, b2 = st.columns(2)
            with b1:
                bulk_ngay     = st.date_input("Ngày ký chung", value=date.today(), format="DD/MM/YYYY", key='bulk_ngay')
            with b2:
                bulk_thoi_han = st.number_input("Thời hạn chung (tháng)", min_value=1, max_value=60, value=12, key='bulk_th')
            if st.button("✅ Áp dụng cho các nhà thầu đã chọn", type="primary", use_container_width=True):
                if selected_bulk:
                    den_ngay_bulk = add_months(bulk_ngay, bulk_thoi_han)
                    for nt in selected_bulk:
                        nha_thau_info[nt] = {
                            'da_ky': True, 'ngay_ky': bulk_ngay,
                            'thoi_han': int(bulk_thoi_han), 'den_ngay': den_ngay_bulk
                        }
                    st.session_state['nha_thau_info'] = nha_thau_info
                    st.success(f"✅ Đã cập nhật {len(selected_bulk)} nhà thầu.")
                    st.rerun()
                else:
                    st.warning("Chưa chọn nhà thầu nào.")

        st.divider()
        search_nt = st.text_input("🔍 Tìm nhà thầu", placeholder="Gõ tên để lọc...", key='search_nt')
        filtered_list = [nt for nt in nha_thau_list if search_nt.lower() in nt.lower()] if search_nt else nha_thau_list

        qa, qb = st.columns(2)
        if qa.button("☑️ Chọn tất cả", use_container_width=True):
            for nt in nha_thau_list: nha_thau_info[nt]['da_ky'] = True
            st.rerun()
        if qb.button("☐ Bỏ chọn tất cả", use_container_width=True):
            for nt in nha_thau_list: nha_thau_info[nt]['da_ky'] = False
            st.rerun()

        st.caption(f"Đang hiển thị {len(filtered_list)}/{len(nha_thau_list)} nhà thầu")
        left_col, right_col = st.columns(2)
        for idx, nt in enumerate(filtered_list):
            info = nha_thau_info.get(nt, {'da_ky': False})
            icon = '✅' if info.get('da_ky') else '⬜'
            with (left_col if idx%2==0 else right_col).expander(f"{icon} {nt}", expanded=False):
                da_ky = st.checkbox("Đã ký phụ lục", key=f"ck_{nt}", value=info.get('da_ky', False))
                if da_ky:
                    i1, i2 = st.columns(2)
                    with i1:
                        ngay_ky = st.date_input("Ngày ký", value=info.get('ngay_ky', date.today()),
                                                 format="DD/MM/YYYY", key=f"ngay_{nt}")
                    with i2:
                        thoi_han = st.number_input("Thời hạn (tháng)", min_value=1, max_value=60,
                                                    value=info.get('thoi_han', 12), key=f"th_{nt}")
                    den_ngay = add_months(ngay_ky, thoi_han)
                    st.caption(f"📅 {ngay_ky.strftime('%d/%m/%Y')} → {den_ngay.strftime('%d/%m/%Y')}")
                    nha_thau_info[nt] = {'da_ky':True,'ngay_ky':ngay_ky,'thoi_han':int(thoi_han),'den_ngay':den_ngay}
                else:
                    nha_thau_info[nt] = {'da_ky': False}

        st.session_state['nha_thau_info'] = nha_thau_info

        def get_ngay(nt, col):
            info = nha_thau_info.get(nt, {})
            if not info.get('da_ky'): return ''
            return to_yyyymmdd(info.get('ngay_ky') if 'TU' in col else info.get('den_ngay'))

        for col_n in ('TU_NGAY_HD','DEN_NGAY_HD','TU_NGAY'):
            df_result[col_n] = df_result['NHA_THAU'].apply(lambda x, c=col_n: get_ngay(x, c))
        df_result['DEN_NGAY'] = ''
        st.session_state['df_result'] = df_result

        da_ky_count = sum(1 for i in nha_thau_info.values() if i.get('da_ky'))
        st.info(f"✅ **{da_ky_count}/{len(nha_thau_list)}** nhà thầu đã ký phụ lục.")

        # ── BƯỚC 7: THỐNG KÊ & XỬ LÝ MÃ ──────────────
        # FIX: Bỏ bảng xem trước (nặng) → chỉ hiện metrics + bảng xử lý mã
        st.markdown('<div class="step-header">📊 Bước 7 — Thống kê & Xử lý mã</div>', unsafe_allow_html=True)

        total   = len(df_result)
        miss_ma = (df_result['MA_THUOC'] == '').sum()
        n_qld   = (df_result.get('IS_QLD_KD', pd.Series()) == 'CẦN RÀ SOÁT').sum()
        n_nodd  = max(0, (df_result['MA_DUONG_DUNG'] == '').sum())
        n_sub   = (df_result.get('IS_MA_CON', pd.Series()) == '1').sum()
        n_multi = (df_result.get('NHIEU_MA', pd.Series()) != '').sum()

        m1, m2, m3, m4, m5, m6 = st.columns(6)
        m1.metric("📊 Tổng",      f"{total:,}")
        m2.metric("✅ Có mã",     f"{total-miss_ma:,}")
        m3.metric("🔴 Thiếu mã",  f"{miss_ma:,}")
        m4.metric("🟠 QLD-KD",    f"{n_qld:,}")
        m5.metric("🟣 Mã .1/.2",  f"{n_sub:,}")
        m6.metric("⚠️ Nhiều mã",  f"{n_multi:,}")

        st.caption("💡 Tải **File Theo dõi** ở Bước 8 để xem toàn bộ dữ liệu bằng Excel với bộ lọc.")

        # ── XỬ LÝ NHIỀU MÃ (NHIEU_MA) ─────────────────
        if n_multi > 0:
            # FIX: build ma_dd_list_map với dedup key = (MA, dd_tight, ten_tight)
            # → giữ cả "Morphin (hydroclorid, sulfat)" và "Morphin" cho cùng MA=40.16, DD=Tiêm
            ma_dd_list_map = {}
            for _, td_r in df_thuoc.iterrows():
                ma_v  = sc(td_r.get('MA',''))
                dd_v  = sc(td_r.get('DUONGDUNG',''))
                ten_v = sc(td_r.get('TEN',''))
                if not ma_v: continue
                ma_dd_list_map.setdefault(ma_v, [])
                # KEY MỚI: (dd_tight, ten_tight) — giữ riêng biệt cả 2 TEN cùng DD
                pair_key = (_norm_tight(dd_v), _norm_tight(ten_v))
                if pair_key not in [(  _norm_tight(p[0]), _norm_tight(p[1])) for p in ma_dd_list_map[ma_v]]:
                    ma_dd_list_map[ma_v].append((dd_v, ten_v))

            with st.expander(
                f"⚠️ **{n_multi} dòng có NHIỀU MÃ cùng đường dùng** — Chọn mã đúng",
                expanded=True
            ):
                multi_rows = df_result[df_result.get('NHIEU_MA','') != ''][
                    ['TEN_HOAT_CHAT','DUONG_DUNG','MA_THUOC','NHIEU_MA']
                ].drop_duplicates(subset=['TEN_HOAT_CHAT','DUONG_DUNG']).reset_index(drop=True)

                for multi_idx, (_, mr) in enumerate(multi_rows.iterrows()):
                    all_mas  = mr['NHIEU_MA'].split(' | ')
                    dd_thau  = mr['DUONG_DUNG']

                    cand_rows  = []
                    seen_pairs = set()
                    for m in all_mas:
                        for dd_v, ten_v in ma_dd_list_map.get(m, [('','')]):
                            # KEY MỚI: (MA, dd_tight, ten_tight)
                            pk = (m, _norm_tight(dd_v), _norm_tight(ten_v))
                            if pk in seen_pairs: continue
                            seen_pairs.add(pk)
                            score = _dd_similarity(dd_thau, dd_v)
                            cand_rows.append((score, m, dd_v, ten_v))
                    cand_rows.sort(key=lambda x: x[0], reverse=True)

                    multi_labels = []
                    for score, m, dd_v, ten_v in cand_rows:
                        lbl = f"[{m}] — {ten_v} — {dd_v}" if dd_v else f"[{m}] — {ten_v}"
                        multi_labels.append(lbl)

                    xc1, xc2 = st.columns([3, 5])
                    with xc1:
                        st.markdown(f"**`{mr['TEN_HOAT_CHAT']}`** | ĐD file thầu: `{mr['DUONG_DUNG']}`")
                    with xc2:
                        if cand_rows:
                            chosen_i = st.selectbox(
                                "Chọn mã",
                                options=list(range(len(cand_rows))),
                                format_func=lambda i, lbs=multi_labels: lbs[i],
                                key=f"sel_ma_{multi_idx}_{mr['TEN_HOAT_CHAT'][:15]}",
                                label_visibility='collapsed'
                            )
                            if st.button("✅ Xác nhận mã này",
                                         key=f"multi_ok_{multi_idx}_{mr['TEN_HOAT_CHAT'][:15]}",
                                         use_container_width=True):
                                _, chosen_ma, chosen_dd, chosen_ten = cand_rows[chosen_i]
                                _df = st.session_state['df_result']
                                mask = (
                                    (_df['TEN_HOAT_CHAT'] == mr['TEN_HOAT_CHAT']) &
                                    (_df['DUONG_DUNG'] == mr['DUONG_DUNG'])
                                )
                                _df.loc[mask, 'MA_THUOC']         = chosen_ma
                                _df.loc[mask, 'TEN_HOAT_CHAT_XK'] = chosen_ten if chosen_ten else mr['TEN_HOAT_CHAT']
                                _df.loc[mask, 'DUONG_DUNG']        = chosen_dd if chosen_dd else mr['DUONG_DUNG']
                                _df.loc[mask, 'MA_DUONG_DUNG']     = lookup_dd(chosen_dd, dd_lk) if chosen_dd else ''
                                _df.loc[mask, 'NHIEU_MA']          = ''
                                st.session_state['df_result'] = _df
                                st.success(f"✅ Đã gán **{chosen_ma}** — {chosen_ten} — {chosen_dd}")
                                st.rerun()
                        else:
                            st.caption("❌ Không tìm thấy thông tin trong file Tân dược")
                    st.divider()

        # ── XỬ LÝ THIẾU MÃ ─────────────────────────────
        if miss_ma > 0 or n_nodd > 0:
            st.markdown('<div class="step-header">✋ Cần xác nhận trước khi xuất file</div>',
                        unsafe_allow_html=True)

        if miss_ma > 0:
            with st.expander(f"🔴 **{miss_ma} dòng thiếu MA_THUOC** — Xem & xử lý ngay", expanded=True):
                miss_rows = df_result[df_result['MA_THUOC']==''][
                    ['TEN_HOAT_CHAT','DUONG_DUNG']
                ].drop_duplicates(subset=['TEN_HOAT_CHAT','DUONG_DUNG']).reset_index(drop=True)

                df_alias_cur = st.session_state['df_alias'].copy()
                alias_changed = False

                def _find_td_candidates(hc_m: str, dd_m: str):
                    hc_tight  = _norm_tight(hc_m)
                    hc_clean  = hc_tight.rstrip('*').strip()
                    n_comp    = _count_components(hc_m)
                    exact_rows, contains_rows = [], []
                    seen_ma_dd = set()

                    for _, td_r in df_thuoc.iterrows():
                        ten_td = sc(td_r.get('TEN', ''))
                        dd_td  = sc(td_r.get('DUONGDUNG', ''))
                        ma_td  = sc(td_r.get('MA', ''))
                        if not ma_td: continue
                        ten_td_t     = _norm_tight(ten_td)
                        ten_td_clean = ten_td_t.rstrip('*').strip()
                        if _count_components(ten_td) != n_comp: continue
                        # FIX: dedup key dùng (MA, dd_tight, ten_tight)
                        pair_key = (ma_td, _norm_tight(dd_td), ten_td_t)
                        if pair_key in seen_ma_dd: continue
                        dd_score = _dd_similarity(dd_m, dd_td)
                        if ten_td_t == hc_tight or ten_td_clean == hc_clean:
                            seen_ma_dd.add(pair_key)
                            exact_rows.append((dd_score, ma_td, dd_td, ten_td, dd_score >= 0.8))
                        elif hc_clean and (hc_clean in ten_td_t or ten_td_clean in hc_tight):
                            seen_ma_dd.add(pair_key)
                            contains_rows.append((dd_score, ma_td, dd_td, ten_td, False))

                    exact_rows.sort(key=lambda x: x[0], reverse=True)
                    contains_rows.sort(key=lambda x: x[0], reverse=True)
                    return exact_rows + contains_rows

                for row_idx, (_, mrow) in enumerate(miss_rows.iterrows()):
                    hc_m   = mrow['TEN_HOAT_CHAT']
                    dd_m   = mrow['DUONG_DUNG']
                    hc_key = f"{hc_m}||{dd_m}"
                    sel_key = f"miss_idx_{row_idx}"

                    td_candidates = _find_td_candidates(hc_m, dd_m)
                    already_alias = (df_alias_cur['TEN_THAU'].str.strip().str.lower()==hc_m.strip().lower()).any()

                    with st.container():
                        rc1, rc2 = st.columns([4, 6])
                        with rc1:
                            st.markdown(f"**🔴 `{hc_m}`**")
                            st.caption(f"ĐD file thầu: `{dd_m or '—'}`")
                        with rc2:
                            if td_candidates:
                                idx_labels = []
                                for i, (dd_score, ma_td, dd_td, ten_td, is_exact) in enumerate(td_candidates):
                                    star = "⭐ " if (is_exact and dd_score >= 0.8) else ""
                                    idx_labels.append(f"{star}[{ma_td}] — {ten_td} — {dd_td}")

                                chosen_idx = st.selectbox(
                                    "Chọn mã",
                                    options=list(range(len(td_candidates))),
                                    format_func=lambda i, lbs=idx_labels: lbs[i],
                                    key=sel_key,
                                    label_visibility='collapsed'
                                )
                                if st.button("✅ Gán mã này", key=f"miss_ok_{row_idx}",
                                             use_container_width=True):
                                    _, chosen_ma, chosen_dd_td, chosen_ten_td, _ = td_candidates[chosen_idx]
                                    _df = st.session_state['df_result']
                                    mask = ((_df['TEN_HOAT_CHAT'] == hc_m) & (_df['DUONG_DUNG'] == dd_m))
                                    _df.loc[mask, 'MA_THUOC']         = chosen_ma
                                    _df.loc[mask, 'TEN_HOAT_CHAT_XK'] = chosen_ten_td
                                    _df.loc[mask, 'DUONG_DUNG']        = chosen_dd_td
                                    _df.loc[mask, 'MA_DUONG_DUNG']     = lookup_dd(chosen_dd_td, dd_lk)
                                    st.session_state['df_result'] = _df
                                    st.session_state['ma_chon_manual'][hc_key] = chosen_ma
                                    st.success(f"✅ Đã gán **{chosen_ma}** — {chosen_ten_td} — {chosen_dd_td}")
                                    st.rerun()
                            else:
                                st.caption("❌ Không tìm thấy trong file Tân dược")
                                if not already_alias:
                                    td_inp = st.text_input(
                                        "Tên đúng trong Tân dược",
                                        key=f"mis_td_{row_idx}_{hc_m[:15]}",
                                        label_visibility='collapsed',
                                        placeholder="Nhập tên chuẩn trong file Tân dược..."
                                    )
                                    if st.button("➕ Lưu alias tên", key=f"mis_add_{row_idx}_{hc_m[:15]}",
                                                 use_container_width=True):
                                        if td_inp.strip():
                                            new_row = pd.DataFrame([{
                                                'TEN_THAU':   hc_m.strip(),
                                                'TEN_TANDUC': td_inp.strip(),
                                                'GHI_CHU':    'Xác nhận trực tiếp'
                                            }])
                                            df_alias_cur = pd.concat([df_alias_cur, new_row], ignore_index=True)
                                            alias_changed = True
                                        else:
                                            st.warning("Chưa nhập tên")
                                else:
                                    st.caption("✅ Đã có alias tên, chạy lại để áp dụng")
                        st.divider()

                if alias_changed:
                    st.session_state['df_alias'] = df_alias_cur
                    st.success("✅ Đã lưu alias. Bấm **'▶️ Chạy xử lý lại'** để áp dụng.")
                    st.rerun()

        # ── XỬ LÝ ĐƯỜNG DÙNG KHÔNG KHỚP ───────────────
        if n_nodd > 0:
            nodd_mask = (df_result['MA_DUONG_DUNG']=='') & (df_result['DUONG_DUNG'].str.strip()!='')
            nodd_rows = df_result[nodd_mask][
                ['TEN_HOAT_CHAT','DUONG_DUNG']
            ].drop_duplicates().reset_index(drop=True)

            with st.expander(
                f"🟡 **{len(nodd_rows)} đường dùng không khớp bảng chuẩn** — Xem & xác nhận",
                expanded=True
            ):
                st.caption("Chọn [Mã] - Tên - ĐD từ Tân dược → bấm 💾 để ánh xạ.")
                df_dd_alias_cur = st.session_state['df_alias_dd'].copy()
                dd_changed = False

                for nodd_idx, (_, nrow) in enumerate(nodd_rows.iterrows()):
                    hc_n     = nrow['TEN_HOAT_CHAT']
                    dd_n     = nrow['DUONG_DUNG']
                    hc_tight = _norm_tight(hc_n)
                    dd_tight = _norm_tight(dd_n)
                    n_comp_n = _count_components(hc_n)

                    dd_already_mask = (
                        (df_dd_alias_cur['TEN_HOAT_CHAT'].apply(_norm_tight)==hc_tight) &
                        (df_dd_alias_cur['DD_GOC'].apply(_norm_tight)==dd_tight)
                    )
                    dd_already   = dd_already_mask.any()
                    dd_chuan_cur = df_dd_alias_cur[dd_already_mask]['DD_CHUAN'].values[0] if dd_already else ''

                    # Tìm ứng viên — FIX: dedup dùng (MA, dd_tight, ten_tight)
                    nodd_exact, nodd_contains = [], []
                    seen_nodd_pairs = set()
                    hc_clean_n = hc_tight.rstrip('*').strip()
                    for _, td_r in df_thuoc.iterrows():
                        ten_td  = sc(td_r.get('TEN',''))
                        dd_td   = sc(td_r.get('DUONGDUNG',''))
                        ma_td   = sc(td_r.get('MA',''))
                        if not ma_td: continue
                        ten_td_t     = _norm_tight(ten_td)
                        ten_td_clean = ten_td_t.rstrip('*').strip()
                        if _count_components(ten_td) != n_comp_n: continue
                        # FIX key
                        pair_key_nodd = (ma_td, _norm_tight(dd_td), ten_td_t)
                        if pair_key_nodd in seen_nodd_pairs: continue
                        dd_score = _dd_similarity(dd_n, dd_td)
                        if ten_td_t == hc_tight or ten_td_clean == hc_clean_n:
                            seen_nodd_pairs.add(pair_key_nodd)
                            nodd_exact.append((dd_score, ma_td, dd_td, ten_td))
                        elif hc_clean_n and (hc_clean_n in ten_td_t or ten_td_clean in hc_tight):
                            seen_nodd_pairs.add(pair_key_nodd)
                            nodd_contains.append((dd_score, ma_td, dd_td, ten_td))
                    nodd_exact.sort(key=lambda x: x[0], reverse=True)
                    nodd_contains.sort(key=lambda x: x[0], reverse=True)
                    td_cands_nodd = nodd_exact + nodd_contains

                    nodd_ma_opts, nodd_ma_labels = [], []
                    for dd_score, ma_td, dd_td, ten_td in td_cands_nodd:
                        star = "⭐ " if dd_score >= 0.8 else ""
                        nodd_ma_opts.append(ma_td)
                        nodd_ma_labels.append(f"{star}[{ma_td}] - {ten_td} - {dd_td}")

                    with st.container():
                        dc1, dc2, dc3 = st.columns([3, 5, 1])
                        with dc1:
                            st.markdown(f"**🟡 `{hc_n}`**")
                            st.caption(f"ĐD file thầu: `{dd_n or '—'}`")
                        with dc2:
                            if dd_already:
                                st.success(f"✅ Đã ánh xạ → **`{dd_chuan_cur}`**")
                                if st.button("🔄 Chọn lại", key=f"dd_reset_{nodd_idx}_{hc_n[:15]}"):
                                    df_dd_alias_cur = df_dd_alias_cur[~dd_already_mask].reset_index(drop=True)
                                    st.session_state['df_alias_dd'] = df_dd_alias_cur
                                    st.rerun()
                            elif nodd_ma_opts:
                                sel_nodd_chosen_idx = st.selectbox(
                                    "Chọn mã",
                                    options=list(range(len(nodd_ma_opts))),
                                    format_func=lambda i, lbs=nodd_ma_labels: lbs[i],
                                    key=f"dd_sel_ma_{nodd_idx}_{hc_n[:15]}",
                                    label_visibility='collapsed',
                                )
                            else:
                                dd_all    = sorted(df_thuoc['DUONGDUNG'].dropna().unique().tolist())
                                sel_dd_raw = st.selectbox(
                                    "Chọn đường dùng chuẩn",
                                    ['-- Chọn đường dùng chuẩn --'] + dd_all,
                                    key=f"dd_fallback_{nodd_idx}_{hc_n[:15]}",
                                    label_visibility='collapsed',
                                )
                        with dc3:
                            if not dd_already:
                                if st.button("💾", key=f"dd_save_{nodd_idx}_{hc_n[:15]}",
                                             use_container_width=True, help="Lưu ánh xạ đường dùng"):
                                    if nodd_ma_opts:
                                        _cand         = td_cands_nodd[sel_nodd_chosen_idx]
                                        dd_chuan_save = _cand[2]
                                        ma_save       = _cand[1]
                                    else:
                                        dd_chuan_save = sel_dd_raw if sel_dd_raw != '-- Chọn đường dùng chuẩn --' else ''
                                        ma_save       = None
                                    if dd_chuan_save:
                                        _df      = st.session_state['df_result']
                                        mask_n   = (_df['TEN_HOAT_CHAT']==hc_n) & (_df['DUONG_DUNG']==dd_n)
                                        if mask_n.any():
                                            _df.loc[mask_n, 'DUONG_DUNG']    = dd_chuan_save.strip()
                                            _df.loc[mask_n, 'MA_DUONG_DUNG'] = lookup_dd(dd_chuan_save, dd_lk)
                                            if nodd_ma_opts and ma_save:
                                                _cand_save     = td_cands_nodd[sel_nodd_chosen_idx]
                                                ten_chuan_save = _cand_save[3]
                                                _df.loc[mask_n, 'MA_THUOC']         = ma_save
                                                _df.loc[mask_n, 'TEN_HOAT_CHAT_XK'] = ten_chuan_save
                                            st.session_state['df_result'] = _df
                                        new_row = pd.DataFrame([{
                                            'TEN_HOAT_CHAT': hc_n.strip(),
                                            'DD_GOC':        dd_n.strip(),
                                            'DD_CHUAN':      dd_chuan_save.strip(),
                                            'GHI_CHU':       'Xác nhận trực tiếp'
                                        }])
                                        df_dd_alias_cur = pd.concat([df_dd_alias_cur, new_row], ignore_index=True)
                                        dd_changed = True
                                        st.success(f"✅ Đã lưu: {hc_n} — {dd_chuan_save}")
                                    else:
                                        st.warning("Chưa chọn")
                        st.divider()

                if dd_changed:
                    st.session_state['df_alias_dd'] = df_dd_alias_cur
                    st.rerun()

        # ── BƯỚC 8: XUẤT FILE ───────────────────────
        st.markdown('<div class="step-header">📥 Bước 8 — Xuất file MAU_03</div>', unsafe_allow_html=True)

        def prepare_export(filter_nt=None):
            df_out = st.session_state['df_result'].copy()
            if filter_nt is not None:
                df_out = df_out[df_out['NHA_THAU'].isin(filter_nt)]
            if 'TEN_HOAT_CHAT_XK' in df_out.columns:
                df_out['TEN_HOAT_CHAT'] = df_out['TEN_HOAT_CHAT_XK']
            for col in MAU03_COLS:
                if col not in df_out.columns: df_out[col] = ''
            return df_out[MAU03_COLS + ['IS_QLD_KD','IS_MA_CON']].reset_index(drop=True)

        mau03_bytes_val = st.session_state['mau03_bytes']
        nha_thau_da_ky  = [nt for nt,info in nha_thau_info.items() if info.get('da_ky')]

        btn1, btn2 = st.columns(2)
        with btn1:
            st.markdown("""
**📌 File Ánh xạ** *(nạp lên cổng BHYT)*
- Chỉ nhà thầu đã ký ✅
- 🔴 Thiếu MA_THUOC | 🟠 QLD-KD | 🟡 Đường dùng lạ | 🟣 Mã .1/.2
            """)
            if nha_thau_da_ky:
                df_ax = prepare_export(filter_nt=nha_thau_da_ky)
                st.download_button(
                    f"⬇️ Tải file Ánh xạ ({len(df_ax):,} dòng / {len(nha_thau_da_ky)} NCC)",
                    export_mau03(df_ax, mau03_bytes_val),
                    f"AnhXa_MAU03_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, type="primary")
            else:
                st.warning("⚠️ Chưa có nhà thầu nào được tick 'Đã ký phụ lục'.")

        with btn2:
            st.markdown("""
**📊 File Theo dõi** *(xem bằng Excel, báo cáo hối thúc)*
- Toàn bộ danh mục kể cả chưa ký
- Dùng bộ lọc Excel để rà soát mã còn thiếu
            """)
            df_td = prepare_export(filter_nt=None)
            st.download_button(
                f"⬇️ Tải file Theo dõi ({len(df_td):,} dòng / {len(nha_thau_list)} NCC)",
                export_mau03(df_td, mau03_bytes_val),
                f"TheoDoi_MAU03_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

        # Bảng nhà thầu
        with st.expander("📋 Bảng tổng hợp nhà thầu", expanded=False):
            rows_nt = [{'Nhà thầu':nt,'Đã ký':'✅' if i.get('da_ky') else '⬜',
                        'Ngày ký':i['ngay_ky'].strftime('%d/%m/%Y') if i.get('ngay_ky') else '',
                        'TU_NGAY_HD':to_yyyymmdd(i.get('ngay_ky')),
                        'Hết hạn':i['den_ngay'].strftime('%d/%m/%Y') if i.get('den_ngay') else '',
                        'DEN_NGAY_HD':to_yyyymmdd(i.get('den_ngay')),
                        'Thời hạn':i.get('thoi_han',''),
                        'Số dòng':int((df_result['NHA_THAU']==nt).sum())}
                       for nt,i in [(nt,nha_thau_info.get(nt,{})) for nt in nha_thau_list]]
            st.dataframe(pd.DataFrame(rows_nt), use_container_width=True, hide_index=True)

        # FIX 6: Bỏ expander "Bảng mã đường dùng chuẩn" (không cần thiết)

# ════════════════════════════════════════════════════
# TAB 2: ALIAS
# ════════════════════════════════════════════════════
with tab_alias:
    st.markdown("""
### ✏️ Quản lý Alias — Ánh xạ tên hoạt chất & đường dùng

Alias giúp hệ thống tự nhận ra những khác biệt giữa file thầu và file Tân dược.
- **Alias Tên**: tên hoạt chất file thầu ≠ tên trong file Tân dược
- **Alias Đường dùng**: đường dùng file thầu không có trong bảng chuẩn BHYT
""")

    subtab_ten, subtab_dd = st.tabs(["📝 Alias Tên hoạt chất", "🔬 Alias Đường dùng"])

    with subtab_ten:
        st.caption("**Mục đích:** Tên file thầu khác tên Tân dược → thêm alias để tự khớp.")
        df_alias = st.session_state['df_alias'].copy()

        st.markdown("#### 📋 Alias hiện có")
        if df_alias.empty:
            st.info("Chưa có alias nào.")
        else:
            st.dataframe(df_alias, use_container_width=True, hide_index=True)
            with st.expander("🗑️ Xóa alias"):
                del_idx = st.number_input("Số thứ tự dòng cần xóa (từ 0)",
                                           min_value=0, max_value=max(0,len(df_alias)-1), value=0)
                if st.button("Xóa dòng"):
                    df_alias = df_alias.drop(index=del_idx).reset_index(drop=True)
                    st.session_state['df_alias'] = df_alias
                    st.rerun()

        st.divider()
        st.markdown("#### ➕ Thêm alias mới")
        a1, a2, a3 = st.columns([3,3,2])
        with a1: inp_thau = st.text_input("Tên trong file THẦU", placeholder="Magnesi aspartat + kali aspartat")
        with a2: inp_td   = st.text_input("Tên ĐÚNG trong file Tân dược", placeholder="Magnesi aspartat+ kali aspartat")
        with a3: inp_note = st.text_input("Ghi chú", placeholder="Lệch dấu cách")

        if st.button("➕ Thêm alias", type="primary", use_container_width=True):
            if not inp_thau.strip() or not inp_td.strip():
                st.error("❌ Phải nhập đủ cả 2 trường tên.")
            elif (df_alias['TEN_THAU'].str.strip().str.lower() == inp_thau.strip().lower()).any():
                st.warning("⚠️ Alias đã tồn tại.")
            else:
                new_row = pd.DataFrame([{'TEN_THAU':inp_thau.strip(),'TEN_TANDUC':inp_td.strip(),
                                          'GHI_CHU':inp_note.strip()}])
                df_alias = pd.concat([df_alias, new_row], ignore_index=True)
                st.session_state['df_alias'] = df_alias
                st.success(f"✅ Đã thêm alias. Bấm 'Chạy xử lý lại' để áp dụng."); st.rerun()

        # Gợi ý từ kết quả
        if st.session_state.get('df_result') is not None:
            df_r = st.session_state['df_result']
            missing_hc = df_r[df_r['MA_THUOC']==''][['TEN_HOAT_CHAT','DUONG_DUNG']].drop_duplicates().reset_index(drop=True)
            if not missing_hc.empty:
                st.divider()
                st.markdown("#### 🔍 Hoạt chất chưa khớp từ lần xử lý gần nhất")
                for _, miss_row in missing_hc.iterrows():
                    hc_miss = miss_row['TEN_HOAT_CHAT']; dd_miss = miss_row['DUONG_DUNG']
                    with st.container():
                        st.markdown(f"**🔴 `{hc_miss}`** | Đường dùng: `{dd_miss or '—'}`")
                        qa2, qb2, qc2 = st.columns([4,2,1])
                        with qa2:
                            td_val = st.text_input("Tên đúng trong Tân dược", value='',
                                                    key=f"ias_td_{hc_miss}_{dd_miss}",
                                                    label_visibility='collapsed',
                                                    placeholder="Nhập tên chuẩn...")
                        with qb2:
                            dd_val = st.text_input("Đường dùng", value=dd_miss,
                                                    key=f"ias_dd_{hc_miss}_{dd_miss}",
                                                    label_visibility='collapsed')
                        with qc2:
                            if st.button("➕", key=f"btn_ias_{hc_miss}_{dd_miss}", use_container_width=True):
                                if td_val.strip():
                                    exists = (df_alias['TEN_THAU'].str.strip().str.lower()==hc_miss.strip().lower()).any()
                                    if not exists:
                                        new_row = pd.DataFrame([{'TEN_THAU':hc_miss.strip(),'TEN_TANDUC':td_val.strip(),
                                                                  'GHI_CHU':'Từ danh sách chưa khớp'}])
                                        df_alias = pd.concat([df_alias, new_row], ignore_index=True)
                                        st.session_state['df_alias'] = df_alias
                                        st.success(f"✅ Đã thêm alias cho **{hc_miss}**"); st.rerun()
                                    else: st.warning("Đã tồn tại.")
                        st.divider()
            else:
                st.success("🎉 Tất cả hoạt chất đã khớp MA_THUOC!")

        st.divider()
        st.markdown("##### 💾 Sao lưu alias_ten.csv")
        dl1, dl2 = st.columns(2)
        with dl1:
            if not df_alias.empty:
                st.download_button("⬇️ Tải alias_ten.csv",
                                   df_alias.to_csv(index=False).encode('utf-8'),
                                   "alias_ten.csv", "text/csv", use_container_width=True)
        with dl2:
            uploaded_ten = st.file_uploader("⬆️ Upload alias_ten.csv", type=['csv'], key='upload_alias_ten')
            if uploaded_ten:
                try:
                    df_up = pd.read_csv(uploaded_ten, dtype=str).fillna('')
                    for col in ALIAS_COLS:
                        if col not in df_up.columns: df_up[col] = ''
                    st.session_state['df_alias'] = df_up[ALIAS_COLS]
                    st.success(f"✅ Đã nạp {len(df_up)} alias tên."); st.rerun()
                except Exception as e:
                    st.error(f"❌ Lỗi: {e}")

    with subtab_dd:
        st.caption("Đường dùng trong file thầu không có trong bảng chuẩn BHYT. Xác nhận → tự nhớ lần sau.")
        df_alias_dd_tab = st.session_state['df_alias_dd'].copy()

        if df_alias_dd_tab.empty:
            st.info("Chưa có alias đường dùng nào.")
        else:
            st.dataframe(df_alias_dd_tab, use_container_width=True, hide_index=True)
            with st.expander("🗑️ Xóa alias đường dùng"):
                del_dd_idx = st.number_input("Số thứ tự dòng cần xóa (từ 0)",
                                              min_value=0, max_value=max(0,len(df_alias_dd_tab)-1), value=0,
                                              key='del_dd_idx')
                if st.button("Xóa dòng alias ĐD", key='del_dd_btn'):
                    df_alias_dd_tab = df_alias_dd_tab.drop(index=del_dd_idx).reset_index(drop=True)
                    st.session_state['df_alias_dd'] = df_alias_dd_tab
                    st.rerun()

        st.divider()
        st.markdown("##### 💾 Sao lưu alias_dd.csv")
        dl3, dl4 = st.columns(2)
        with dl3:
            if not df_alias_dd_tab.empty:
                st.download_button("⬇️ Tải alias_dd.csv",
                                   df_alias_dd_tab.to_csv(index=False).encode('utf-8'),
                                   "alias_dd.csv", "text/csv", use_container_width=True)
        with dl4:
            uploaded_dd = st.file_uploader("⬆️ Upload alias_dd.csv", type=['csv'], key='upload_alias_dd')
            if uploaded_dd:
                try:
                    df_up_dd = pd.read_csv(uploaded_dd, dtype=str).fillna('')
                    for col in ALIAS_DD_COLS:
                        if col not in df_up_dd.columns: df_up_dd[col] = ''
                    st.session_state['df_alias_dd'] = df_up_dd[ALIAS_DD_COLS]
                    st.success(f"✅ Đã nạp {len(df_up_dd)} alias đường dùng."); st.rerun()
                except Exception as e:
                    st.error(f"❌ Lỗi: {e}")
